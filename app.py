import streamlit as st
import pandas as pd
import numpy as np
import io
import os
import base64
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from matching_engine import reconcile_dfs
from sample_generator import generate_sample_data
from pdf_extractor import extract_bank_statement_pdf

# Configuration de la page Streamlit
st.set_page_config(
    page_title="MAN Truck - Rapprochement Bancaire",
    page_icon="🚛",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- INITIALISATION DE L'ÉTAT D'AUTHENTIFICATION ---
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

# Écran de connexion si non authentifié
if not st.session_state.authenticated:
    st.markdown("""
        <style>
        /* Masquer la barre latérale sur l'écran de connexion */
        [data-testid="stSidebar"] {
            display: none;
        }
        /* Ajuster le conteneur principal de Streamlit pour le centrer */
        .main .block-container {
            max-width: 480px !important;
            padding-top: 8rem !important;
            padding-bottom: 5rem !important;
            margin: auto !important;
        }
        /* Arrière-plan dégradé moderne bleu nuit / ardoise */
        .stApp {
            background: linear-gradient(135deg, #0F172A 0%, #1E293B 100%) !important;
        }
        /* Styliser le formulaire Streamlit sous forme de carte blanche premium */
        div[data-testid="stForm"] {
            background-color: #FFFFFF !important;
            border-radius: 16px !important;
            border: 1px solid #E2E8F0 !important;
            padding: 2.5rem !important;
            box-shadow: 0 25px 50px -12px rgba(0, 0, 0, 0.25) !important;
        }
        /* Cacher la bordure par défaut de Streamlit */
        div[data-testid="stForm"] > div {
            border: none !important;
            padding: 0 !important;
        }
        </style>
    """, unsafe_allow_html=True)
    
    with st.container():
        # Charger le logo localement et l'encoder en base64 pour l'intégrer proprement à l'HTML
        logo_html = ""
        logo_path = "logo.png"
        if os.path.exists(logo_path):
            try:
                with open(logo_path, "rb") as f:
                    logo_base64 = base64.b64encode(f.read()).decode()
                logo_html = f'<img src="data:image/png;base64,{logo_base64}" style="width: 140px; margin-bottom: 1.5rem;" alt="MAN Logo">'
            except Exception:
                pass
        
        # Fallback si le logo n'est pas trouvé
        if not logo_html:
            logo_html = '<div style="font-size: 3.5rem; margin-bottom: 1.5rem;">🚛</div>'
            
        with st.form("login_form", clear_on_submit=False):
            st.markdown(f"""
                <div style="text-align: center;">
                    {logo_html}
                    <h2 style="font-size: 1.6rem; font-weight: 700; color: #0F172A; margin: 0 0 0.5rem 0; font-family: 'Plus Jakarta Sans', sans-serif;">Accès Sécurisé</h2>
                    <p style="font-size: 0.85rem; color: #64748B; margin: 0 0 2rem 0; line-height: 1.4; font-family: 'Plus Jakarta Sans', sans-serif;">
                        Veuillez saisir votre code d'accès pour accéder à la plateforme de rapprochement bancaire MAN Truck.
                    </p>
                </div>
            """, unsafe_allow_html=True)
            
            # Saisie et validation du code
            code_input = st.text_input("Code d'accès", type="password", placeholder="Saisir le code d'accès...")
            submit = st.form_submit_button("Se connecter 🔓", use_container_width=True, type="primary")
            
            if submit:
                if code_input == "MAN2026":
                    st.session_state.authenticated = True
                    st.toast("Connexion réussie !", icon="✅")
                    st.rerun()
                else:
                    st.error("Code d'accès incorrect. Veuillez réessayer.")
                    
    st.stop() # Arrêter le rendu du reste de la page


# Thème CSS personnalisé premium (Intégration d'un style moderne, polices Google Fonts et design épuré)
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@300;400;500;600;700&display=swap');
    
    html, body, [class*="css"] {
        font-family: 'Plus Jakarta Sans', sans-serif;
    }
    
    /* En-tête principal de la page */
    .main-header {
        background: linear-gradient(135deg, #0F172A 0%, #1E293B 100%);
        padding: 2.5rem;
        border-radius: 16px;
        color: white;
        margin-bottom: 2rem;
        box-shadow: 0 10px 15px -3px rgba(0, 0, 0, 0.1), 0 4px 6px -2px rgba(0, 0, 0, 0.05);
        position: relative;
        overflow: hidden;
    }
    .main-header::before {
        content: "";
        position: absolute;
        top: -50%;
        right: -20%;
        width: 300px;
        height: 300px;
        background: rgba(59, 130, 246, 0.1);
        border-radius: 50%;
        filter: blur(50px);
    }
    
    .main-header h1 {
        margin: 0;
        font-weight: 700;
        font-size: 2.2rem;
        letter-spacing: -0.025em;
        color: #F8FAFC;
    }
    .main-header p {
        margin-top: 0.5rem;
        margin-bottom: 0;
        font-weight: 400;
        font-size: 1.1rem;
        color: #94A3B8;
    }
    
    /* Cartes Métriques Premium */
    .metric-card {
        background: #FFFFFF;
        border: 1px solid #E2E8F0;
        border-radius: 12px;
        padding: 1.25rem;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);
        transition: transform 0.2s ease, box-shadow 0.2s ease;
    }
    .metric-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 10px 15px -3px rgba(0, 0, 0, 0.05);
    }
    .metric-title {
        font-size: 0.85rem;
        font-weight: 600;
        color: #64748B;
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }
    .metric-value {
        font-size: 1.75rem;
        font-weight: 700;
        margin-top: 0.25rem;
    }
    
    /* Couleur thématique */
    .text-success { color: #10B981 !important; }
    .text-danger { color: #EF4444 !important; }
    .text-warning { color: #F59E0B !important; }
    .text-primary { color: #3B82F6 !important; }
    
    /* Séparateurs et styles */
    hr {
        margin: 2rem 0;
        border-color: #E2E8F0;
    }
    </style>
""", unsafe_allow_html=True)


def to_excel_premium(df_a, df_b, df_c):
    """
    Génère un fichier Excel stylisé de qualité premium en mémoire avec openpyxl.
    """
    wb = Workbook()
    
    # Supprimer la feuille par défaut créée par openpyxl
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    sheets_info = [
        ("Tableau A - Matches", df_a, "3B82F6"), # Bleu pour les matches
        ("Tableau B - Écarts Compta", df_b, "EF4444"), # Rouge pour les écarts compta
        ("Tableau C - Écarts Banque", df_c, "F59E0B") # Ambre pour les écarts banque
    ]
    
    for sheet_name, df, theme_color in sheets_info:
        ws = wb.create_sheet(title=sheet_name)
        ws.views.sheetView[0].showGridLines = True
        
        # Styles openpyxl
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color=theme_color, end_color=theme_color, fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_font = Font(name='Segoe UI', size=10)
        border_side = Side(border_style='thin', color='CBD5E1')
        data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        # Écriture du titre de l'onglet
        ws.merge_cells("A1:I1")
        ws["A1"] = f"Rapport Rapprochement : {sheet_name}"
        ws["A1"].font = Font(name='Segoe UI', size=14, bold=True, color='1E293B')
        ws["A1"].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 35
        
        # Laisser la ligne 2 vide
        
        # Convertir et ajouter le dataframe à partir de la ligne 3
        df_display = df.copy()
        
        # Écriture des en-têtes
        headers = df_display.columns.tolist()
        for col_idx, header_name in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = data_border
        ws.row_dimensions[3].height = 28
        
        # Écriture des données
        for r_idx, row in enumerate(dataframe_to_rows(df_display, index=False, header=False), start=4):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = data_font
                cell.border = data_border
                
                # Formatage spécifique des nombres
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                elif isinstance(val, (pd.Timestamp, np.datetime64)):
                    cell.number_format = 'yyyy-mm-dd'
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='left')
            ws.row_dimensions[r_idx].height = 20
            
        # Ajustement de la taille des colonnes
        for col in ws.columns:
            max_len = 0
            col_letter = None
            for cell in col:
                if col_letter is None and hasattr(cell, 'column_letter'):
                    col_letter = cell.column_letter
                # Ne pas prendre en compte la première ligne fusionnée pour le calcul de taille
                if cell.row == 1:
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            if col_letter:
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    # Sauvegarde dans un buffer
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# --- HEADER DE L'APPLICATION ---
st.markdown("""
    <div class="main-header">
        <h1>🚛 Plateforme Intelligente de Rapprochement Bancaire</h1>
        <p>Rapprochez instantanément votre Grand Livre comptable avec votre Relevé Bancaire grâce à des algorithmes multi-critères avancés.</p>
    </div>
""", unsafe_allow_html=True)


# --- INITIALISATION DE L'ÉTAT DE SESSION ---
if 'compta_df' not in st.session_state:
    st.session_state.compta_df = None
if 'banque_df' not in st.session_state:
    st.session_state.banque_df = None
if 'demo_loaded' not in st.session_state:
    st.session_state.demo_loaded = False


# --- SIDEBAR DE CONFIGURATION ---
st.sidebar.image("logo.png", width=120)
st.sidebar.markdown("### 🚛 Rapprochement Bancaire")
st.sidebar.info("Cet outil permet d'automatiser le rapprochement entre le Grand Livre et le Relevé Bancaire en se basant sur les numéros de chèques, les remises et les proximités temporelles.")

date_tolerance = 10


# --- ONGLETS PRINCIPAUX ---
tab_pdf, tab_import, tab_matching = st.tabs(["📄 Convertisseur PDF en Excel", "📂 1. Importation & Correspondance", "⚡ 2. Rapprochement & Analyses"])

with tab_pdf:
    st.markdown("### 📄 Convertisseur intelligent de Relevés PDF en Excel")
    st.write("Téléchargez un relevé bancaire PDF (ex: Attijariwafa bank) contenant plusieurs pages, et laissez Gemini 2.5 Flash extraire toutes les transactions sous forme de tableau Excel propre.")
    
    uploaded_pdf = st.file_uploader("Déposer le relevé bancaire PDF", type=["pdf"], key="uploaded_pdf_uploader")
    
    if uploaded_pdf is not None:
        st.info(f"Fichier chargé : `{uploaded_pdf.name}` ({len(uploaded_pdf.getvalue()) / 1024:.1f} Ko)")
        
        if st.button("⚡ Convertir en Excel avec Gemini 2.5", type="primary", use_container_width=True):
            with st.spinner("Analyse du relevé et extraction des transactions en cours (cela peut prendre de 1 à 2 minutes)..."):
                try:
                    pdf_bytes = uploaded_pdf.getvalue()
                    df_extracted = extract_bank_statement_pdf(pdf_bytes)
                    
                    st.success(f"Extraction réussie ! **{len(df_extracted)}** transactions ont été extraites avec succès.")
                    
                    # Affichage des métriques de contrôle
                    sum_deb = df_extracted['debit'].fillna(0).sum()
                    sum_cre = df_extracted['credit'].fillna(0).sum()
                    
                    col_st1, col_st2 = st.columns(2)
                    with col_st1:
                        st.metric("Total Débits Extraits", f"{sum_deb:,.2f} DH")
                    with col_st2:
                        st.metric("Total Crédits Extraits", f"{sum_cre:,.2f} DH")
                        
                    # Aperçu
                    st.markdown("#### 👀 Aperçu des données extraites")
                    st.dataframe(df_extracted, use_container_width=True)
                    
                    # Génération Excel
                    excel_output = io.BytesIO()
                    with pd.ExcelWriter(excel_output, engine='openpyxl') as writer:
                        df_extracted.to_excel(writer, index=False, sheet_name="Relevé Extrait")
                    excel_bytes = excel_output.getvalue()
                    
                    st.download_button(
                        label="📥 Télécharger le relevé au format Excel",
                        data=excel_bytes,
                        file_name=f"{os.path.splitext(uploaded_pdf.name)[0]}_extrait.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True
                    )
                    
                except Exception as e:
                    st.error(f"Une erreur est survenue lors de l'extraction : {str(e)}")

with tab_import:
    st.markdown("### 📥 Importation des données comptables et bancaires")
    st.write("Choisissez votre méthode d'importation puis chargez vos données.")
    
    import_mode = st.radio(
        "Méthode d'importation :",
        ["Deux fichiers Excel séparés", "Un seul fichier Excel (contenant les 2 feuilles)"],
        horizontal=True,
        key="import_mode_selector"
    )
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    if import_mode == "Deux fichiers Excel séparés":
        col_u1, col_u2 = st.columns(2)
        
        with col_u1:
            st.markdown("#### **Grand Livre (Comptabilité)**")
            uploaded_compta = st.file_uploader("Fichier Grand Livre Excel", type=["xlsx", "xls"], key="u_compta")
            if uploaded_compta is not None:
                st.session_state.compta_df = pd.read_excel(uploaded_compta)
                st.session_state.demo_loaded = False
                st.toast("Grand Livre importé !", icon="✅")
                
        with col_u2:
            st.markdown("#### **Relevé Bancaire**")
            uploaded_banque = st.file_uploader("Fichier Relevé Bancaire Excel", type=["xlsx", "xls"], key="u_banque")
            if uploaded_banque is not None:
                st.session_state.banque_df = pd.read_excel(uploaded_banque)
                st.session_state.demo_loaded = False
                st.toast("Relevé bancaire importé !", icon="✅")
    else:
        uploaded_single = st.file_uploader("Fichier Excel Unique (Multi-feuilles)", type=["xlsx", "xls"], key="u_single")
        if uploaded_single is not None:
            try:
                xls = pd.ExcelFile(uploaded_single)
                sheet_names = xls.sheet_names
                
                col_s1, col_s2 = st.columns(2)
                with col_s1:
                    # Auto-détecter l'onglet compta par mots-clés
                    def_c_idx = next((i for i, s in enumerate(sheet_names) if 'compta' in s.lower() or 'livre' in s.lower() or 'grand' in s.lower()), 0)
                    compta_sheet = st.selectbox("Sélectionnez la feuille du **Grand Livre**", sheet_names, index=def_c_idx)
                with col_s2:
                    # Auto-détecter l'onglet banque par mots-clés
                    def_b_idx = next((i for i, s in enumerate(sheet_names) if 'banq' in s.lower() or 'rele' in s.lower() or 'rlv' in s.lower() or 'statement' in s.lower()), min(1, len(sheet_names)-1))
                    banque_sheet = st.selectbox("Sélectionnez la feuille du **Relevé Bancaire**", sheet_names, index=def_b_idx)
                
                if compta_sheet == banque_sheet:
                    st.warning("⚠️ Attention, vous avez sélectionné la même feuille pour la Comptabilité et la Banque.")
                    
                st.session_state.compta_df = pd.read_excel(uploaded_single, sheet_name=compta_sheet)
                st.session_state.banque_df = pd.read_excel(uploaded_single, sheet_name=banque_sheet)
                st.session_state.demo_loaded = False
                st.toast("Les deux feuilles ont été importées avec succès !", icon="✅")
                
            except Exception as e:
                st.error(f"Erreur lors de l'analyse du fichier Excel : {str(e)}")

    # --- CONFIGURATION DU MAPPING ---
    if st.session_state.compta_df is not None and st.session_state.banque_df is not None:
        st.markdown("<hr>", unsafe_allow_html=True)
        
        c_cols = st.session_state.compta_df.columns.tolist()
        b_cols = st.session_state.banque_df.columns.tolist()
        
        # Fonction interne de recherche intelligente tolérante aux accents
        def find_smart_col(cols, keywords, default_val=None):
            for keyword in keywords:
                for col in cols:
                    col_norm = col.lower().replace('é', 'e').replace('è', 'e').replace('ê', 'e').replace('à', 'a').replace('â', 'a').replace('î', 'i').replace('ï', 'i').replace('ç', 'c')
                    if keyword in col_norm:
                        return col
            return default_val
            
        # Détection automatique des colonnes standards
        c_date_val = find_smart_col(c_cols, ['date ecriture', 'date comptable', 'date op', 'date'], None)
        c_lib_val = find_smart_col(c_cols, ['libelle', 'description', 'desc'], None)
        c_deb_val = find_smart_col(c_cols, ['debit'], None)
        c_cre_val = find_smart_col(c_cols, ['credit'], None)
        
        b_date_val = find_smart_col(b_cols, ['date comptable', 'date valeur', 'date op', 'date'], None)
        b_lib_val = find_smart_col(b_cols, ['libelle', 'description', 'desc', 'texte'], None)
        b_deb_val = find_smart_col(b_cols, ['debit'], None)
        b_cre_val = find_smart_col(b_cols, ['credit'], None)
        
        # Si on a trouvé les colonnes vitales, on auto-map
        if (c_date_val and c_lib_val and c_deb_val and c_cre_val and 
            b_date_val and b_lib_val and b_deb_val and b_cre_val):
            
            st.success("✅ Les colonnes standards ont été détectées automatiquement. Aucun mapping manuel n'est requis.")
            
            st.session_state.compta_mapping = {
                'date': c_date_val, 'ref': None, 'libelle': c_lib_val,
                'debit': c_deb_val, 'credit': c_cre_val, 'montant': None
            }
            st.session_state.banque_mapping = {
                'date': b_date_val, 'ref': None, 'libelle': b_lib_val,
                'debit': b_deb_val, 'credit': b_cre_val, 'montant': None
            }
            
        else:
            st.markdown("### 🗺️ Configuration de la correspondance des colonnes (Mapping)")
            st.info("Les colonnes standards n'ont pas pu être entièrement détectées. Veuillez les lier manuellement.")
            
            col_m1, col_m2 = st.columns(2)
            
            with col_m1:
                st.markdown("<div style='background-color:#F8FAFC; padding:1.25rem; border-radius:10px; border-left: 4px solid #3B82F6;'>", unsafe_allow_html=True)
                st.markdown("**📁 MAPPING GRAND LIVRE (COMPTA)**")
                
                c_date = st.selectbox("Colonne Date", c_cols, index=c_cols.index(c_date_val) if c_date_val in c_cols else 0)
                
                c_ref_val = find_smart_col(c_cols, ['piece', 'ref', 'cheq', 'num'], None)
                c_ref_options = [None] + c_cols
                c_ref = st.selectbox("Colonne Référence / Chèque", c_ref_options, index=c_ref_options.index(c_ref_val) if c_ref_val in c_ref_options else 0)
                
                c_lib_options = [None] + c_cols
                c_lib = st.selectbox("Colonne Libellé", c_lib_options, index=c_lib_options.index(c_lib_val) if c_lib_val in c_lib_options else 0)
                
                c_amt_mode_index = 0 if (c_deb_val and c_cre_val) else 1
                c_amount_mode = st.radio("Structure des montants (Compta)", ["Débit / Crédit séparés", "Montant unique (+/-)"], index=c_amt_mode_index, key="c_amt_mode")
                
                if c_amount_mode == "Débit / Crédit séparés":
                    c_debit = st.selectbox("Colonne Débit", c_cols, index=c_cols.index(c_deb_val) if c_deb_val in c_cols else 0)
                    c_credit = st.selectbox("Colonne Crédit", c_cols, index=c_cols.index(c_cre_val) if c_cre_val in c_cols else 0)
                    c_montant = None
                else:
                    c_amt_val = find_smart_col(c_cols, ['montant', 'valeur', 'solde', 'amount'], c_cols[0])
                    c_montant = st.selectbox("Colonne Montant", c_cols, index=c_cols.index(c_amt_val) if c_amt_val in c_cols else 0)
                    c_debit = c_credit = None
                st.markdown("</div>", unsafe_allow_html=True)
                
            with col_m2:
                st.markdown("<div style='background-color:#F8FAFC; padding:1.25rem; border-radius:10px; border-left: 4px solid #F59E0B;'>", unsafe_allow_html=True)
                st.markdown("**📁 MAPPING RELEVÉ BANCAIRE**")
                
                b_date = st.selectbox("Colonne Date", b_cols, index=b_cols.index(b_date_val) if b_date_val in b_cols else 0)
                
                b_ref_val = find_smart_col(b_cols, ['piece', 'ref', 'cheq', 'num'], None)
                b_ref_options = [None] + b_cols
                b_ref = st.selectbox("Colonne Référence / Chèque", b_ref_options, index=b_ref_options.index(b_ref_val) if b_ref_val in b_ref_options else 0)
                
                b_lib_options = [None] + b_cols
                b_lib = st.selectbox("Colonne Libellé", b_lib_options, index=b_lib_options.index(b_lib_val) if b_lib_val in b_lib_options else 0)
                
                b_amt_mode_index = 0 if (b_deb_val and b_cre_val) else 1
                b_amount_mode = st.radio("Structure des montants (Banque)", ["Débit / Crédit séparés", "Montant unique (+/-)"], index=b_amt_mode_index, key="b_amt_mode")
                
                if b_amount_mode == "Débit / Crédit séparés":
                    b_debit = st.selectbox("Colonne Débit", b_cols, index=b_cols.index(b_deb_val) if b_deb_val in b_cols else 0)
                    b_credit = st.selectbox("Colonne Crédit", b_cols, index=b_cols.index(b_cre_val) if b_cre_val in b_cols else 0)
                    b_montant = None
                else:
                    b_amt_val = find_smart_col(b_cols, ['montant', 'valeur', 'solde', 'amount'], b_cols[0])
                    b_montant = st.selectbox("Colonne Montant", b_cols, index=b_cols.index(b_amt_val) if b_amt_val in b_cols else 0)
                    b_debit = b_credit = None
                st.markdown("</div>", unsafe_allow_html=True)
                
            # Enregistrement des mappings dans l'état de session
            st.session_state.compta_mapping = {
                'date': c_date, 'ref': c_ref, 'libelle': c_lib,
                'debit': c_debit, 'credit': c_credit, 'montant': c_montant
            }
            st.session_state.banque_mapping = {
                'date': b_date, 'ref': b_ref, 'libelle': b_lib,
                'debit': b_debit, 'credit': b_credit, 'montant': b_montant
            }
        
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown("### 👀 Aperçu rapide des fichiers importés")
        
        col_p1, col_p2 = st.columns(2)
        with col_p1:
            st.markdown("**Aperçu Grand Livre (5 premières lignes) :**")
            st.dataframe(st.session_state.compta_df.head(5), use_container_width=True)
        with col_p2:
            st.markdown("**Aperçu Relevé Bancaire (5 premières lignes) :**")
            st.dataframe(st.session_state.banque_df.head(5), use_container_width=True)
            
    else:
        # Aucun fichier encore chargé
        st.markdown("<div style='text-align: center; padding: 3rem; background-color: #F8FAFC; border-radius: 12px; margin-top:2rem;'>", unsafe_allow_html=True)
        st.markdown("### 🫙 En attente d'importation de fichiers")
        st.markdown("Veuillez importer vos fichiers Excel (Grand Livre et Relevé Bancaire) pour commencer le rapprochement.")
        st.markdown("</div>", unsafe_allow_html=True)


with tab_matching:
    if st.session_state.compta_df is not None and st.session_state.banque_df is not None:
        st.markdown("### ⚡ Moteur d'analyse du Rapprochement Bancaire")
        st.write("Cliquez sur le bouton ci-dessous pour lancer l'analyse complète de réconciliation multiniveau.")
        
        if st.button("🚀 Lancer le Rapprochement automatique", type="primary", use_container_width=True):
            with st.spinner("Exécution des algorithmes de matching avancés..."):
                try:
                    df_a, df_b, df_c = reconcile_dfs(
                        st.session_state.compta_df,
                        st.session_state.banque_df,
                        st.session_state.compta_mapping,
                        st.session_state.banque_mapping,
                        date_tolerance=date_tolerance
                    )
                    
                    # 1. CALCULS DES INDICATEURS CLÉS (METRICS)
                    total_compta_lines = len(st.session_state.compta_df)
                    total_banque_lines = len(st.session_state.banque_df)
                    matched_lines_count = len(df_a)
                    
                    # Nombre de lignes uniques d'origines qui ont été matchées
                    # (Pour la remise, plusieurs compta lines pointent vers 1 bank line)
                    matched_compta_count = total_compta_lines - len(df_b)
                    matched_banque_count = total_banque_lines - len(df_c)
                    
                    reconciliation_rate = (matched_compta_count / total_compta_lines) * 100 if total_compta_lines > 0 else 0
                    
                    # 2. AFFICHAGE DES MÉTRIQUES PREMIUM EN GRILLE
                    st.markdown("#### 📊 Résumé des indicateurs")
                    
                    m_col1, m_col2, m_col3, m_col4 = st.columns(4)
                    
                    with m_col1:
                        st.markdown(f"""
                            <div class="metric-card">
                                <div class="metric-title">🎯 Taux de Rapprochement</div>
                                <div class="metric-value text-success">{reconciliation_rate:.1f} %</div>
                            </div>
                        """, unsafe_allow_html=True)
                        
                    with m_col2:
                        st.markdown(f"""
                            <div class="metric-card">
                                <div class="metric-title">✅ Écritures Réconciliées</div>
                                <div class="metric-value text-primary">{matched_lines_count}</div>
                            </div>
                        """, unsafe_allow_html=True)
                        
                    with m_col3:
                        st.markdown(f"""
                            <div class="metric-card">
                                <div class="metric-title">🔴 Écarts Comptabilité</div>
                                <div class="metric-value text-danger">{len(df_b)}</div>
                            </div>
                        """, unsafe_allow_html=True)
                        
                    with m_col4:
                        st.markdown(f"""
                            <div class="metric-card">
                                <div class="metric-title">🟡 Écarts Banque</div>
                                <div class="metric-value text-warning">{len(df_c)}</div>
                            </div>
                        """, unsafe_allow_html=True)
                        
                    # Graphique Plotly de répartition
                    st.markdown("<br>", unsafe_allow_html=True)
                    fig = go.Figure(data=[go.Pie(
                        labels=['Écritures Comptabilité Réconciliées', 'Écarts Comptabilité', 'Écarts Banque'],
                        values=[matched_compta_count, len(df_b), len(df_c)],
                        hole=.4,
                        marker_colors=['#10B981', '#EF4444', '#F59E0B']
                    )])
                    fig.update_layout(
                        title_text="Visualisation des correspondances et des écarts",
                        title_x=0.0,
                        margin=dict(t=40, b=10, l=10, r=10),
                        height=350,
                        legend=dict(orientation="h", yanchor="bottom", y=-0.1, xanchor="center", x=0.5)
                    )
                    st.plotly_chart(fig, use_container_width=True)
                    
                    st.markdown("<hr>", unsafe_allow_html=True)
                    st.markdown("### 📋 Rapports et Tableaux détaillés")
                    
                    # 3. AFFICHAGE DES TROIS FEUILLES DU RAPPORT
                    tab_a_view, tab_b_view, tab_c_view = st.tabs([
                        "🟢 Tableau A (Matches)", 
                        "🔴 Tableau B (Écarts Compta)", 
                        "🟡 Tableau C (Écarts Banque)"
                    ])
                    
                    with tab_a_view:
                        st.markdown("##### **Liste des écritures rapprochées (Tableau A)**")
                        st.caption("Ce tableau liste les écritures dont la correspondance a été trouvée automatiquement par référence, par remise ou par montant identique avec proximité temporelle.")
                        st.dataframe(df_a, use_container_width=True)
                        
                    with tab_b_view:
                        st.markdown("##### **Écritures présentes en comptabilité non trouvées en banque (Tableau B)**")
                        st.caption("Ces écritures sont présentes dans votre comptabilité mais n'apparaissent pas sur le relevé bancaire (ex: chèques émis non encore encaissés).")
                        st.dataframe(df_b, use_container_width=True)
                        
                    with tab_c_view:
                        st.markdown("##### **Écritures du relevé bancaire non saisies en comptabilité (Tableau C)**")
                        st.caption("Ces transactions sont apparues sur votre relevé bancaire mais n'ont pas d'équivalent dans votre Grand Livre (ex: agios, frais, virements reçus inconnus).")
                        st.dataframe(df_c, use_container_width=True)
                        
                    # 4. EXPORT EXCEL PREMIUM
                    st.markdown("<hr>", unsafe_allow_html=True)
                    st.markdown("### 💾 Exportation du Rapport Final")
                    st.write("Téléchargez le rapport complet de rapprochement bancaire au format Excel avec 3 onglets distincts mis en forme de manière professionnelle.")
                    
                    excel_data = to_excel_premium(df_a, df_b, df_c)
                    
                    st.download_button(
                        label="📥 Télécharger le rapport Excel de Rapprochement",
                        data=excel_data,
                        file_name="rapport_rapprochement_bancaire_final.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True
                    )
                    
                except Exception as e:
                    st.error(f"Une erreur est survenue lors de l'exécution du rapprochement bancaire : {str(e)}")
                    st.info("Veuillez vérifier que le mapping de vos colonnes est correct et correspond bien aux données contenues dans vos fichiers.")
                    
    else:
        st.markdown("<div style='text-align: center; padding: 3rem; background-color: #F8FAFC; border-radius: 12px; margin-top:2rem;'>", unsafe_allow_html=True)
        st.markdown("### 🫙 Aucune donnée chargée")
        st.markdown("Veuillez d'abord importer vos fichiers Excel dans le premier onglet **📂 1. Importation & Correspondance**.")
        st.markdown("</div>", unsafe_allow_html=True)
